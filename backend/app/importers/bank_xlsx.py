from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Any
import warnings

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

PARSER_VERSION = "bank-xlsx-v1"
MAPPING_VERSION = "preview-rules-v1"

EXPECTED_HEADERS = {
    "date": "Дата",
    "bank_category_raw": "Категорія",
    "payment_instrument_label": "Картка",
    "description_raw": "Опис операції",
    "amount": "Сума в валюті картки",
    "currency": "Валюта картки",
    "transaction_amount": "Сума в валюті транзакції",
    "transaction_currency": "Валюта транзакції",
    "balance_after": "Залишок на кінець періоду",
    "balance_currency": "Валюта залишку",
}


class BankXlsxParseError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedBankOperation:
    row_number: int
    status: str
    reason_codes: list[str]
    occurred_at: datetime | None
    amount: Decimal | None
    currency: str | None
    transaction_amount: Decimal | None
    transaction_currency: str | None
    balance_after: Decimal | None
    payment_instrument_label: str | None
    bank_category_raw: str | None
    description_raw: str | None
    merchant_name: str | None
    proposed_flow_type: str | None
    proposed_scope: str | None
    confidence: Decimal | None
    error_message: str | None
    normalized_payload: dict[str, Any]


@dataclass(frozen=True)
class BankStatementSummary:
    period_start: datetime | None
    period_end: datetime | None
    total_rows: int
    auto_ready_count: int
    needs_review_count: int
    imported_count: int
    excluded_count: int
    duplicate_count: int
    error_count: int
    uncategorized_count: int
    work_fop_count: int
    savings_count: int
    parser_version: str
    mapping_version: str


@dataclass(frozen=True)
class BankStatementParseResult:
    source_filename: str
    sheet_name: str
    summary: BankStatementSummary
    rows: list[ParsedBankOperation]


def parse_bank_xlsx(path: str | Path) -> BankStatementParseResult:
    source_path = Path(path)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]

        header_row_number, header_positions = _find_header_row(worksheet)
        period_start, period_end = _parse_period(_as_str(worksheet.cell(row=1, column=1).value))

        parsed_rows: list[ParsedBankOperation] = []
        for row in worksheet.iter_rows(min_row=header_row_number + 1, values_only=False):
            values_by_key = {
                key: row[column_index - 1].value
                for key, column_index in header_positions.items()
                if column_index <= len(row)
            }
            if not any(value is not None for value in values_by_key.values()):
                continue
            parsed_rows.append(_parse_operation(row[0].row, values_by_key))

        summary = _build_summary(period_start, period_end, parsed_rows)
        return BankStatementParseResult(
            source_filename=source_path.name,
            sheet_name=worksheet.title,
            summary=summary,
            rows=parsed_rows,
        )
    finally:
        workbook.close()


def _find_header_row(worksheet: Worksheet) -> tuple[int, dict[str, int]]:
    expected_values = set(EXPECTED_HEADERS.values())

    for row in worksheet.iter_rows(min_row=1, max_row=20, values_only=False):
        labels = {_as_str(cell.value): cell.column for cell in row if _as_str(cell.value)}
        if expected_values.issubset(labels):
            return row[0].row, {
                key: labels[header]
                for key, header in EXPECTED_HEADERS.items()
            }

    raise BankXlsxParseError("Could not find expected bank XLSX headers.")


def _parse_operation(row_number: int, values: dict[str, Any]) -> ParsedBankOperation:
    error_messages: list[str] = []
    occurred_at = _parse_datetime(values.get("date"), error_messages)
    amount = _parse_decimal(values.get("amount"), "amount", error_messages)
    transaction_amount = _parse_decimal(values.get("transaction_amount"), "transaction_amount", error_messages)
    balance_after = _parse_decimal(values.get("balance_after"), "balance_after", error_messages)

    bank_category_raw = _as_str(values.get("bank_category_raw"))
    description_raw = _as_str(values.get("description_raw"))
    payment_instrument_label = _as_str(values.get("payment_instrument_label"))
    currency = _as_str(values.get("currency"))
    transaction_currency = _as_str(values.get("transaction_currency"))

    proposed_flow_type = _infer_flow_type(bank_category_raw, description_raw, amount)
    proposed_scope = _infer_scope(bank_category_raw, description_raw)
    reason_codes = _reason_codes(bank_category_raw, description_raw, amount, proposed_flow_type, proposed_scope)
    confidence = Decimal("0.9000") if not reason_codes else Decimal("0.6000")

    status = "error" if error_messages else "needs_review" if reason_codes else "auto_ready"

    return ParsedBankOperation(
        row_number=row_number,
        status=status,
        reason_codes=reason_codes if status != "error" else [*reason_codes, "parse_error"],
        occurred_at=occurred_at,
        amount=amount,
        currency=currency,
        transaction_amount=transaction_amount,
        transaction_currency=transaction_currency,
        balance_after=balance_after,
        payment_instrument_label=payment_instrument_label,
        bank_category_raw=bank_category_raw,
        description_raw=description_raw,
        merchant_name=_extract_merchant_name(description_raw),
        proposed_flow_type=proposed_flow_type,
        proposed_scope=proposed_scope,
        confidence=confidence if status != "error" else None,
        error_message="; ".join(error_messages) or None,
        normalized_payload={
            "date_raw": _as_str(values.get("date")),
            "balance_currency": _as_str(values.get("balance_currency")),
            "direction": _infer_direction(amount),
            "duplicate_key": _duplicate_key(occurred_at, amount, description_raw, payment_instrument_label),
        },
    )


def _build_summary(
    period_start: datetime | None,
    period_end: datetime | None,
    rows: list[ParsedBankOperation],
) -> BankStatementSummary:
    return BankStatementSummary(
        period_start=period_start,
        period_end=period_end,
        total_rows=len(rows),
        auto_ready_count=sum(row.status == "auto_ready" for row in rows),
        needs_review_count=sum(row.status == "needs_review" for row in rows),
        imported_count=sum(row.status != "excluded" and row.status != "error" for row in rows),
        excluded_count=sum(row.status == "excluded" for row in rows),
        duplicate_count=sum("duplicate" in row.reason_codes for row in rows),
        error_count=sum(row.status == "error" for row in rows),
        uncategorized_count=sum("uncategorized" in row.reason_codes for row in rows),
        work_fop_count=sum(row.proposed_scope == "work_fop" for row in rows),
        savings_count=sum(row.proposed_flow_type == "transfer_to_savings" for row in rows),
        parser_version=PARSER_VERSION,
        mapping_version=MAPPING_VERSION,
    )


def _parse_period(text: str | None) -> tuple[datetime | None, datetime | None]:
    if not text:
        return None, None
    matches = re.findall(r"\d{2}\.\d{2}\.\d{4}", text)
    if len(matches) < 2:
        return None, None
    return (
        datetime.strptime(matches[0], "%d.%m.%Y"),
        datetime.strptime(matches[1], "%d.%m.%Y"),
    )


def _parse_datetime(value: Any, errors: list[str]) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = _as_str(value)
    if not text:
        errors.append("date is empty")
        return None
    try:
        return datetime.strptime(text, "%d.%m.%Y %H:%M:%S")
    except ValueError:
        errors.append(f"invalid date: {text}")
        return None


def _parse_decimal(value: Any, field_name: str, errors: list[str]) -> Decimal | None:
    if value is None:
        errors.append(f"{field_name} is empty")
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        errors.append(f"invalid {field_name}: {value}")
        return None


def _infer_direction(amount: Decimal | None) -> str | None:
    if amount is None:
        return None
    if amount < 0:
        return "expense"
    if amount > 0:
        return "income"
    return "transfer"


def _infer_flow_type(
    bank_category_raw: str | None,
    description_raw: str | None,
    amount: Decimal | None,
) -> str | None:
    category = _lower(bank_category_raw)
    description = _lower(description_raw)

    if "скарбнич" in description or "заощад" in category:
        return "transfer_to_savings"
    if "зняття готівки" in category:
        return "cash_withdrawal"
    if "переказ на свою" in category or "зарахування зі своєї" in category:
        return "transfer_to_own_account"
    if "переказ" in category:
        return "person_transfer"
    if "реквізит" in category:
        return "requisites_payment"
    if amount is not None and amount > 0:
        return "income"
    if amount is not None and amount < 0:
        return "purchase"
    return None


def _infer_scope(bank_category_raw: str | None, description_raw: str | None) -> str:
    text = f"{_lower(bank_category_raw)} {_lower(description_raw)}"
    if any(marker in text for marker in ("fop", "фоп", "github", "openai", "upwork")):
        return "work_fop"
    return "family"


def _reason_codes(
    bank_category_raw: str | None,
    description_raw: str | None,
    amount: Decimal | None,
    proposed_flow_type: str | None,
    proposed_scope: str,
) -> list[str]:
    codes: list[str] = []
    category = _lower(bank_category_raw)

    if proposed_flow_type == "person_transfer":
        codes.append("person_transfer")
    if proposed_flow_type == "requisites_payment":
        codes.append("requisites_payment")
    if proposed_scope == "work_fop":
        codes.append("work_fop_candidate")
    if amount is not None and abs(amount) >= Decimal("5000.00"):
        codes.append("large_amount")
    if "супермаркет" in category and amount is not None and abs(amount) >= Decimal("2500.00"):
        codes.append("large_supermarket")
    if not bank_category_raw and not description_raw:
        codes.append("uncategorized")

    return codes


def _extract_merchant_name(description_raw: str | None) -> str | None:
    text = _as_str(description_raw)
    if not text:
        return None
    return re.sub(r"\s+", " ", text).strip()


def _duplicate_key(
    occurred_at: datetime | None,
    amount: Decimal | None,
    description_raw: str | None,
    payment_instrument_label: str | None,
) -> str | None:
    if occurred_at is None or amount is None:
        return None
    parts = [
        occurred_at.isoformat(),
        str(amount),
        _lower(description_raw),
        _lower(payment_instrument_label),
    ]
    return "|".join(parts)


def _lower(value: str | None) -> str:
    return (value or "").casefold()


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
